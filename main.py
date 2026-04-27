#!/usr/bin/env python3
import logging, re, random, os
from datetime import datetime
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BOT_TOKEN = "8650205673:AAFHrBrUaisrXuqwU9Ww60o54zWOkp2POyk"
DATA_FILE = "participants.xlsx"

logging.basicConfig(level=logging.INFO)

COMMENTS = [
    "🔥 Дуже хочу виграти! Це просто неймовірно!",
    "✨ Беру участь! Удачі всім учасникам 🍀",
    "💫 Обожнюю такі розіграші! Тримаю кулачки 🤞",
    "🎉 Беру участь з великим задоволенням!",
    "🌟 Вже мрію про перемогу! Дякую за розіграш!",
    "❤️ Чудовий розіграш! Беру участь!",
    "🏆 Дуже хочеться виграти! Беру участь!",
    "💎 Неймовірний приз! Дуже хочу виграти!",
    "🌈 Беру участь! Нехай виграє найщасливіший!",
    "🚀 Підписаний давно, дуже хочу виграти! 🙏",
    "💝 Дякую за можливість! Беру участь!",
    "⭐ Обожнюю ваш акаунт! Беру участь у розіграші!",
    "🎁 Такий класний приз! Дуже хочу виграти!",
    "🦋 Беру участь! Нехай фортуна посміхнеться мені!",
    "🍀 Удачі всім! Але особливо мені 😄",
    "🥳 Це саме те що я шукав! Беру участь!",
    "💪 Завжди мріяв про такий приз! Беру участь!",
    "🌸 Дякую за розіграш! Дуже хочу виграти!",
    "🎯 Мета — перемога! Беру участь! 🏅",
]

def thin_border():
    s = Side(style="thin", color="D8B4FE")
    return Border(left=s, right=s, top=s, bottom=s)

def create_excel(participants):
    wb = Workbook()
    ws = wb.active
    ws.title = "Учасники"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "🎁 Учасники розіграшу"
    c.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="7B2FBE")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Згенеровано: {datetime.now().strftime('%d.%m.%Y о %H:%M')}   |   Всього: {len(participants)} учасників"
    c.font = Font(name="Arial", size=10, italic=True, color="7B2FBE")
    c.fill = PatternFill("solid", fgColor="EDE9FE")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6

    for col, h in enumerate(["№", "Нік", "Коментар", "Дата додавання"], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C084FC")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[4].height = 26

    for i, p in enumerate(participants):
        row = 5 + i
        fill_color = "F5F0FF" if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate(p[:4], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10, color="1E0A2E")
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal="center" if col in (1, 4) else "left",
                vertical="center", wrap_text=(col == 3)
            )
        ws.row_dimensions[row].height = 22

    if not participants:
        ws.merge_cells("A5:D5")
        c = ws.cell(row=5, column=1, value="Учасників ще немає")
        c.font = Font(name="Arial", size=11, italic=True, color="9CA3AF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 30
    else:
        fr = 5 + len(participants)
        ws.merge_cells(f"A{fr}:D{fr}")
        c = ws.cell(row=fr, column=1, value=f"Загалом учасників: {len(participants)}")
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7B2FBE")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fr].height = 24

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 20
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def load_participants():
    if not os.path.exists(DATA_FILE):
        return []
    try:
        wb = load_workbook(DATA_FILE, data_only=True)
        ws = wb.active
        result = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is not None and row[1] is not None:
                result.append(list(row[:4]))
        return result
    except Exception:
        return []

def save_participants(participants):
    buf = create_excel(participants)
    with open(DATA_FILE, "wb") as f:
        f.write(buf.read())

def main_menu_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Додати учасників", callback_data="add"),
         InlineKeyboardButton("📋 Список", callback_data="list")],
        [InlineKeyboardButton("📥 Скачати Excel", callback_data="download"),
         InlineKeyboardButton("🏆 Вибрати переможця", callback_data="winner")],
        [InlineKeyboardButton("💬 Коментар для мене", callback_data="comment"),
         InlineKeyboardButton("📊 Статистика", callback_data="stats")],
        [InlineKeyboardButton("🗑 Очистити список", callback_data="clear")],
    ])

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    p = load_participants()
    await update.message.reply_text(
        f"🎁 *Giveaway Bot*\n\n👥 Учасників зараз: *{len(p)}*\n\n"
        "✅ Зберігаю ніки учасників\n✅ Автоматично нумерую\n"
        "✅ Надсилаю красивий Excel файл\n✅ Вибираю випадкового переможця\n"
        "✅ Генерую коментарі для участі\n\nОбери дію 👇",
        parse_mode="Markdown", reply_markup=main_menu_kb()
    )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    d = q.data

    if d == "add":
        context.user_data["mode"] = "add"
        await q.message.reply_text(
            "📝 *Надішли ніки учасників:*\n\n"
            "`@нік коментар`\n`@нік2 інший коментар`\n\n"
            "Або просто ніки — кожен з нового рядка",
            parse_mode="Markdown"
        )
    elif d == "list":
        p = load_participants()
        if not p:
            await q.message.reply_text("📋 Список порожній!")
            return
        lines = [f"`{x[0]}.` {x[1]}" for x in p[-30:]]
        text = f"📋 *Учасники ({len(p)} чол.):*\n\n" + "\n".join(lines)
        if len(p) > 30:
            text += f"\n\n_...та ще {len(p)-30}_"
        await q.message.reply_text(text, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="menu")]]))
    elif d == "download":
        p = load_participants()
        buf = create_excel(p)
        fname = f"rozigrysh_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        await q.message.reply_document(document=buf, filename=fname,
            caption=f"📊 Таблиця учасників\n👥 Всього: {len(p)} чол.")
    elif d == "winner":
        p = load_participants()
        if not p:
            await q.message.reply_text("😕 Немає учасників!")
            return
        w = random.choice(p)
        await q.message.reply_text(
            f"🎉 *ПЕРЕМОЖЕЦЬ!* 🎉\n\n🏆 Учасник *#{w[0]}*\n👤 {w[1]}\n\nВітаємо! 🥳🎁",
            parse_mode="Markdown"
        )
    elif d == "comment":
        c = random.choice(COMMENTS)
        await q.message.reply_text(f"💬 *Коментар для участі:*\n\n{c}", parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🔄 Інший", callback_data="comment"),
                InlineKeyboardButton("🏠 Меню", callback_data="menu")
            ]]))
    elif d == "stats":
        p = load_participants()
        await q.message.reply_text(
            f"📊 *Статистика:*\n\n👥 Учасників: *{len(p)}*\n📅 {datetime.now().strftime('%d.%m.%Y')}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="menu")]]))
    elif d == "clear":
        await q.message.reply_text("⚠️ *Очистити весь список?*", parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Так", callback_data="confirm_clear"),
                InlineKeyboardButton("❌ Ні", callback_data="menu")
            ]]))
    elif d == "confirm_clear":
        save_participants([])
        await q.message.reply_text("✅ Список очищено! Готово до нового розіграшу 🎁",
            reply_markup=main_menu_kb())
    elif d == "menu":
        p = load_participants()
        await q.message.reply_text(
            f"🏠 *Головне меню*\n👥 Учасників: *{len(p)}*",
            parse_mode="Markdown", reply_markup=main_menu_kb()
        )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    participants = load_participants()
    added, skipped = [], []
    for line in lines:
        m = re.match(r"(@\w+)(.*)", line)
        if m:
            nick = m.group(1)
            comment = m.group(2).strip() or "—"
            num = len(participants) + 1
            date = datetime.now().strftime("%d.%m.%Y %H:%M")
            participants.append([num, nick, comment, date])
            added.append(f"`{num}.` {nick}")
        else:
            skipped.append(line[:25])
    if added:
        save_participants(participants)
        result = f"✅ *Додано {len(added)} учасників:*\n\n" + "\n".join(added)
        if skipped:
            result += f"\n\n⚠️ Пропущено {len(skipped)} (без @ніку)"
        await update.message.reply_text(result, parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("📥 Скачати Excel", callback_data="download"),
                InlineKeyboardButton("🏠 Меню", callback_data="menu")
            ]]))
    else:
        await update.message.reply_text("❌ Не знайдено @ніків.\n\nФормат: `@нік коментар`",
            parse_mode="Markdown")

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    print("🤖 Бот запущено!")
    app.run_polling()

if __name__ == "__main__":
    main()
